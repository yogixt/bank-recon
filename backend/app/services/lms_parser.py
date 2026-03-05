"""LMS .xlsx parser with embedded JSON extraction for Bank rows."""

import json
import logging
from typing import Generator

from openpyxl import load_workbook

from app.core.constants import EXCEL_BATCH_SIZE

logger = logging.getLogger(__name__)

# Expected column order (0-indexed)
LMS_COLUMNS = [
    "UserId", "MobileNumber", "RoleName", "Point", "Amount",
    "CreatedOn", "Description", "LastUpdatedOn", "TransID",
    "WithdrawType", "StateName",
]


def _parse_bank_description(description: str) -> dict | None:
    """Parse embedded JSON from Bank row description.

    Format: 'ACCEPTED INQUIRY DATA - {"TOTAL_NUM_RECORDS":...}'
    Returns parsed fields dict or None on failure.
    """
    try:
        if "ACCEPTED INQUIRY DATA - " not in description:
            return None
        json_str = description.split("ACCEPTED INQUIRY DATA - ", 1)[1]
        data = json.loads(json_str)
        rec = data["ALL_RECORDS"][0]
        return {
            "payment_ref_no": rec["PAYMENTREFNO"].upper(),
            "txn_status": rec.get("TXN_STATUS", ""),
            "utr_no": rec.get("UTR_NO", ""),
            "od_amount": float(rec.get("OD_AMOUNT", 0)),
            "bene_name": rec.get("BENE_NAME", ""),
            "ifsc_code": rec.get("IFSC_CODE", ""),
            "credit_acc_no": rec.get("CREDIT_ACC_NO", ""),
            "reference_no": rec.get("REFERENCE_NO", ""),
            "txn_reference_no": rec.get("TXN_REFERENCE_NO", ""),
        }
    except Exception as e:
        logger.warning(f"Failed to parse Bank description JSON: {e}")
        return None


def parse_lms_file(file_path: str) -> Generator[list[dict], None, None]:
    """Parse LMS .xlsx file. Yields batches of entry dicts.

    For Bank/bank rows, parses embedded JSON to extract payment_ref_no etc.
    For TDS/Gift rows, stores raw data without JSON parsing.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows()

    # Find header row
    col_map: dict[str, int] = {}
    for row in rows:
        values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
        # Check for known LMS column names
        if "TransID" in values or "trans_id" in [v.lower().replace(" ", "_") for v in values]:
            for idx, val in enumerate(values):
                col_map[val.strip()] = idx
            break

    if not col_map:
        # Assume default column order
        for idx, name in enumerate(LMS_COLUMNS):
            col_map[name] = idx

    batch: list[dict] = []

    for row in rows:
        cells = [cell.value for cell in row]
        if not cells or len(cells) < 9:
            continue

        # Get column values with fallbacks
        def _get(name: str, default=None):
            idx = col_map.get(name)
            if idx is not None and idx < len(cells):
                return cells[idx]
            return default

        trans_id = str(_get("TransID", "") or "").strip()
        if not trans_id:
            continue

        withdraw_type = str(_get("WithdrawType", "") or "").strip()
        description = str(_get("Description", "") or "")
        amount = 0.0
        try:
            amount = float(_get("Amount", 0) or 0)
        except (ValueError, TypeError):
            pass

        point = 0.0
        try:
            point = float(_get("Point", 0) or 0)
        except (ValueError, TypeError):
            pass

        user_id = None
        try:
            uid = _get("UserId")
            if uid is not None:
                user_id = int(float(uid))
        except (ValueError, TypeError):
            pass

        entry = {
            "user_id": user_id,
            "mobile_number": str(_get("MobileNumber", "") or "").strip(),
            "role_name": str(_get("RoleName", "") or "").strip(),
            "point": point,
            "amount": amount,
            "created_on": _get("CreatedOn"),
            "description": description,
            "last_updated_on": _get("LastUpdatedOn"),
            "trans_id": trans_id,
            "withdraw_type": withdraw_type,
            "state_name": str(_get("StateName", "") or "").strip(),
            # JSON-parsed fields (Bank rows only)
            "payment_ref_no": None,
            "txn_status": None,
            "utr_no": None,
            "bene_name": None,
            "ifsc_code": None,
            "credit_acc_no": None,
            "od_amount": None,
            "reference_no": None,
            "txn_reference_no": None,
        }

        # Parse JSON for Bank/bank rows
        if withdraw_type.strip().lower() == "bank" and description:
            parsed = _parse_bank_description(description)
            if parsed:
                entry.update(parsed)

        batch.append(entry)
        if len(batch) >= EXCEL_BATCH_SIZE:
            yield batch
            batch = []

    if batch:
        yield batch
    wb.close()
