"""Streaming file parsers that avoid loading entire files into RAM."""

import csv
import io
from typing import Generator

from openpyxl import load_workbook

from app.core.constants import EXCEL_BATCH_SIZE


class StreamingExcelParser:
    """Parse Excel bank statement using openpyxl read_only mode.

    Yields batches of dicts with extracted bank entry fields.
    """

    @staticmethod
    def extract_bank_id(description: str) -> str | None:
        """Extract bank ID from description: split on ' - ', take 4th part."""
        try:
            parts = str(description).split(" - ")
            if len(parts) > 3:
                return parts[3].strip().upper()
        except Exception:
            pass
        return None

    @staticmethod
    def parse_amount(value) -> float:
        if value is None:
            return 0.0
        try:
            return float(str(value).replace(",", ""))
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def extract_customer(description: str) -> str:
        parts = str(description).split(" - ")
        if len(parts) > 5:
            return parts[-1].strip()
        return "N/A"

    def parse(self, file_path: str) -> Generator[list[dict], None, None]:
        """Yield batches of bank entry dicts."""
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Find header row
        header_row_idx = None
        col_map: dict[str, int] = {}
        rows = ws.iter_rows()

        for row_idx, row in enumerate(rows):
            values = [str(cell.value) if cell.value is not None else "" for cell in row]
            if "DATE" in values:
                header_row_idx = row_idx
                for col_idx, val in enumerate(values):
                    col_map[val.strip().upper()] = col_idx
                break

        if header_row_idx is None:
            wb.close()
            raise ValueError("Could not find header row (looking for 'DATE' column)")

        batch: list[dict] = []

        for row in rows:  # continues from after the header
            cells = [cell.value for cell in row]
            desc = str(cells[col_map.get("DESCRIPTION", 0)] or "")
            bank_id = self.extract_bank_id(desc)
            if bank_id is None:
                continue

            debit = self.parse_amount(cells[col_map.get("DEBITS", -1)] if "DEBITS" in col_map else None)
            credit = self.parse_amount(cells[col_map.get("CREDITS", -1)] if "CREDITS" in col_map else None)

            entry = {
                "bank_id": bank_id,
                "date": str(cells[col_map.get("DATE", 0)] or ""),
                "description": desc,
                "debit_amount": debit,
                "credit_amount": credit,
                "branch": str(cells[col_map.get("BRANCH", -1)] or "") if "BRANCH" in col_map else "",
                "reference_no": str(cells[col_map.get("REFERENCE NO", -1)] or "") if "REFERENCE NO" in col_map else "",
                "customer_name": self.extract_customer(desc),
            }
            batch.append(entry)
            if len(batch) >= EXCEL_BATCH_SIZE:
                yield batch
                batch = []

        if batch:
            yield batch
        wb.close()


class BridgeFileParser:
    """Parse bridge file: alternating lines (txn_id, bank_id)."""

    @staticmethod
    def parse(file_path: str) -> dict[str, str]:
        """Return {transaction_id: bank_id} mapping."""
        bridge_map: dict[str, str] = {}
        with open(file_path, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f if line.strip()]
        for i in range(0, len(lines), 2):
            if i + 1 < len(lines):
                txn_id = lines[i].strip().upper()
                bank_id = lines[i + 1].strip().upper()
                bridge_map[txn_id] = bank_id
        return bridge_map


class TransactionIdParser:
    """Parse transaction IDs from CSV or plain text."""

    @staticmethod
    def parse(file_path: str) -> list[str]:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()

        if "," in content or "\t" in content:
            # CSV format - take first column
            reader = csv.reader(io.StringIO(content))
            ids = []
            for row in reader:
                if row:
                    val = row[0].strip().upper()
                    if val and val != "TRANSACTION_ID" and val != "ID":
                        ids.append(val)
            return list(set(ids))
        else:
            # Plain text, one per line
            ids = [line.strip().upper() for line in content.split("\n") if line.strip()]
            return list(set(ids))
